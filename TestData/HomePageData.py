import openpyxl


class HomePageData:
    test_HomePage_data = [{"name": "Rahul Shetty ", "email": "rahulshettyacademy@gmail.com", "password": "123456",
                           "gender": "Male"}, {"name": "Vartika Shetty", "email": "hello@gmail.com",
                                               "password": "654321", "gender": "Female"}]

    @staticmethod  # For directly calling method without creating object by Classname.methodname
    def getTestData(test_case_name):   # self parameter is req when you have declared method non-static
        Dict = {}
        book = openpyxl.load_workbook("C:\\Users\\HP\\Documents\\PythonDemo.xlsx")
        sheet = book.active
        for i in range(1, sheet.max_row + 1):  # to get rows
            if sheet.cell(row=i, column=1).value == test_case_name:
                for j in range(2, sheet.max_column + 1):  # to get columns
                    Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value

        return [Dict]
