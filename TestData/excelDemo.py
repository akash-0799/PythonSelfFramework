import openpyxl

book = openpyxl.load_workbook("C:\\Users\\HP\\Documents\\PythonDemo.xlsx")
sheet = book.active
Dict ={}
cell = sheet.cell(row=1, column=2)  # Reading
print(cell.value)

sheet.cell(row=2, column=2).value = "Rahul"  # Writing
print(sheet.cell(row=2, column=2).value)
book.save("PythonDemo.xlsx")
print(sheet.max_row)
print(sheet.max_column)

print(sheet['A5'].value)
print(sheet['B2'].value)

print("********************************")
for i in range(1, sheet.max_row+1):  # to get rows
    if sheet.cell(row=i,column=1).value == "Testcase2":
        for j in range(2, sheet.max_column+1):  # to get columns
            Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value


print(Dict)